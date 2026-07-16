"""
reports_service.py — Report Generation Service
================================================
Generates PDF, Excel, and CSV reports from the authenticated user's real data.
Sources: dashboard metrics, feature data, model info, AI recommendations.
No demo data — all content comes from UserPaths-isolated files.

PDF format: uses fpdf2 (real PDF binary)
Excel format: uses openpyxl
CSV format: plain comma-separated text
"""
from __future__ import annotations

import csv
import json
import logging
import uuid
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

BACKEND_DIR = Path(__file__).resolve().parent.parent.parent
REPORTS_DIR = BACKEND_DIR / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_user_reports_dir(user_id: int) -> Path:
    d = REPORTS_DIR / str(user_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _load_dashboard_data(paths) -> dict[str, Any]:
    try:
        from app.services import dashboard_service
        return dashboard_service.get_dashboard_data(paths)
    except Exception as exc:
        logger.warning("Reports: dashboard unavailable: %s", exc)
        return {}


def _load_recommendations(paths) -> dict[str, Any]:
    try:
        from app.services import recommendation_service
        return recommendation_service.generate_recommendations(paths=paths)
    except Exception as exc:
        logger.warning("Reports: recommendations unavailable: %s", exc)
        return {}


def _load_model_info(paths) -> dict[str, Any]:
    try:
        if paths.model_info_file.exists():
            with open(paths.model_info_file, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as exc:
        logger.warning("Reports: model info unavailable: %s", exc)
    return {}


def _get_chat_history(user_id: int) -> list[dict[str, str]]:
    try:
        from app.database import SessionLocal
        from app.models.history import ChatConversation, ChatMessage
        db = SessionLocal()
        try:
            conv = db.query(ChatConversation).filter(ChatConversation.user_id == user_id).first()
            if conv and conv.messages:
                return [{"role": m.role, "message": m.message} for m in conv.messages]
        finally:
            db.close()
    except Exception as exc:
        logger.warning("Reports service: failed to load chat context: %s", exc)
    return []


def _get_latest_scenarios(user_id: int) -> list[dict[str, Any]]:
    try:
        from app.database import SessionLocal
        from app.models.history import ScenarioHistory
        db = SessionLocal()
        try:
            runs = db.query(ScenarioHistory).filter(ScenarioHistory.user_id == user_id).order_by(ScenarioHistory.created_at.desc()).limit(3).all()
            return [{
                "name": r.scenario_name,
                "type": r.scenario_type,
                "revenue": r.predicted_revenue,
                "roas": r.predicted_roas,
                "profit": r.estimated_profit,
                "date": r.created_at.strftime("%Y-%m-%d")
            } for r in runs]
        finally:
            db.close()
    except Exception as exc:
        logger.warning("Reports service: failed to load scenario history context: %s", exc)
    return []


# ---------------------------------------------------------------------------
# PDF Generator (fpdf2)
# ---------------------------------------------------------------------------

def _safe_str(text: str) -> str:
    """Strip characters outside latin-1 range so fpdf2 Helvetica does not crash."""
    if not text:
        return ""
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _generate_pdf(filepath: Path, dashboard: dict, recs: dict, model_info: dict, user_id: int) -> None:
    """Generate a styled PDF report using fpdf2."""
    from fpdf import FPDF

    summary = dashboard.get("summary", {})
    channels = dashboard.get("channel_spend_distribution", [])
    recommendations = recs.get("recommendations", [])
    monthly = dashboard.get("monthly_revenue", [])
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # --- Header ---
    pdf.set_fill_color(15, 23, 42)   # dark bg
    pdf.rect(0, 0, 210, 32, style="F")
    pdf.set_text_color(34, 211, 238)  # brand cyan
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_y(8)
    pdf.cell(0, 10, "ForecastIQ", align="C", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(180, 190, 210)
    pdf.cell(0, 8, "Enterprise Performance Report", align="C", ln=True)

    # --- Metadata bar ---
    pdf.set_fill_color(30, 41, 59)
    pdf.rect(0, 32, 210, 10, style="F")
    pdf.set_y(34)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(130, 140, 160)
    pdf.cell(105, 6, _safe_str(f"  Generated: {now}"), align="L")
    pdf.cell(105, 6, "Confidential - Internal Use Only  ", align="R", ln=True)

    pdf.set_y(46)
    pdf.set_text_color(15, 23, 42)

    # --- Helper functions ---
    def section_header(title: str) -> None:
        pdf.set_y(pdf.get_y() + 4)
        pdf.set_fill_color(34, 211, 238)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 8, _safe_str(f"  {title}"), fill=True, ln=True)
        pdf.set_text_color(30, 41, 59)
        pdf.set_y(pdf.get_y() + 2)

    def kv_row(key: str, value: str, shade: bool = False) -> None:
        if shade:
            pdf.set_fill_color(245, 247, 250)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(80, 90, 110)
        pdf.cell(80, 7, _safe_str(f"  {key}"), fill=True)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(110, 7, _safe_str(str(value)), fill=True, ln=True)

    # --- Executive Summary ---
    section_header("EXECUTIVE SUMMARY")

    metrics = [
        ("Total Revenue", f"${summary.get('total_revenue', 0):,.2f}"),
        ("Total Spend", f"${summary.get('total_spend', 0):,.2f}"),
        ("Overall Profit", f"${summary.get('overall_profit', 0):,.2f}"),
        ("Profit Margin", f"{summary.get('profit_margin', 0):.2f}%"),
        ("Average ROAS", f"{summary.get('average_roas', 0):.2f}x"),
        ("Average CTR", f"{summary.get('average_ctr', 0):.2f}%"),
        ("Avg Daily Revenue", f"${summary.get('average_daily_revenue', 0):,.2f}"),
        ("Avg Daily Spend", f"${summary.get('average_daily_spend', 0):,.2f}"),
    ]
    for i, (k, v) in enumerate(metrics):
        kv_row(k, v, shade=(i % 2 == 0))

    # --- Channel Performance ---
    if channels:
        section_header("CHANNEL SPEND DISTRIBUTION")
        # Table header
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(70, 7, "  Channel", fill=True)
        pdf.cell(50, 7, "Spend ($)", fill=True, align="R")
        pdf.cell(40, 7, "Share (%)", fill=True, align="R")
        pdf.cell(30, 7, "", fill=True, ln=True)
        # Rows
        for i, ch in enumerate(channels):
            shade = (i % 2 == 0)
            if shade:
                pdf.set_fill_color(245, 247, 250)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(30, 41, 59)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(70, 7, _safe_str(f"  {ch.get('channel', '')}"), fill=True)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(50, 7, f"${ch.get('spend', 0):>12,.0f}", fill=True, align="R")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(40, 7, f"{ch.get('percentage', 0):.1f}%", fill=True, align="R")
            pdf.cell(30, 7, "", fill=True, ln=True)

    # --- Model Performance ---
    section_header("MODEL PERFORMANCE")
    if model_info:
        model_metrics = [
            ("Model Name", model_info.get("model_name", "LightGBM")),
            ("R2 Score", f"{model_info.get('r2_score', model_info.get('test_r2', 0)):.4f}"),
            ("Mean Absolute Error", f"${model_info.get('mae', 0):,.2f}"),
            ("Training Date", str(model_info.get("trained_at", "N/A"))),
        ]
        for i, (k, v) in enumerate(model_metrics):
            kv_row(k, str(v), shade=(i % 2 == 0))
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(120, 130, 150)
        pdf.cell(0, 8, "  No model trained yet.", ln=True)

    # --- AI Recommendations ---
    if recommendations:
        section_header("AI RECOMMENDATIONS")
        for i, rec in enumerate(recommendations[:6], 1):
            priority = rec.get("priority", "Medium")
            title = rec.get("title", f"Recommendation {i}")
            desc = (rec.get("description") or "")[:180]
            color = {"High": (239, 68, 68), "Medium": (245, 158, 11), "Low": (34, 197, 94)}.get(priority, (100, 100, 100))
            # Priority badge
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_y(20)
            pdf.set_y(pdf.get_y() + 2)
            pdf.set_fill_color(*color)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            pdf.cell(20, 5, _safe_str(f" {priority.upper()} "), fill=True, align="C")
            pdf.set_text_color(15, 23, 42)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 5, _safe_str(f"  {title}"), ln=True)
            if desc:
                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(80, 90, 110)
                pdf.set_x(pdf.get_x() + 22)
                pdf.multi_cell(0, 5, _safe_str(desc))
            pdf.set_y(pdf.get_y() + 1)

    # --- Monthly Revenue (simple table) ---
    if monthly:
        section_header("MONTHLY REVENUE TREND")
        per_row = 6
        month_chunks = [monthly[i:i+per_row] for i in range(0, len(monthly), per_row)]
        for chunk in month_chunks:
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 8)
            col_w = 190 // len(chunk)
            for m in chunk:
                pdf.cell(col_w, 7, _safe_str(m.get("month", "")), fill=True, align="C")
            pdf.ln()
            pdf.set_fill_color(245, 247, 250)
            pdf.set_text_color(15, 23, 42)
            pdf.set_font("Helvetica", "", 8)
            for m in chunk:
                pdf.cell(col_w, 7, f"${m.get('revenue', 0):,.0f}", fill=True, align="C")
            pdf.ln()
            pdf.set_y(pdf.get_y() + 3)

    # --- Persistent Scenario History ---
    scenarios = _get_latest_scenarios(user_id)
    if scenarios:
        if pdf.get_y() > 220:
            pdf.add_page()
            pdf.set_y(20)
        section_header("LATEST SAVED BUDGET SIMULATIONS & SCENARIOS")
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(60, 7, "  Scenario Name", fill=True)
        pdf.cell(30, 7, "Type", fill=True, align="C")
        pdf.cell(35, 7, "Proj. Revenue ($)", fill=True, align="R")
        pdf.cell(35, 7, "Proj. ROAS", fill=True, align="R")
        pdf.cell(30, 7, "Saved Date", fill=True, align="C", ln=True)

        for i, s in enumerate(scenarios):
            shade = (i % 2 == 0)
            pdf.set_fill_color(245, 247, 250 if shade else 255)
            pdf.set_text_color(30, 41, 59)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(60, 7, _safe_str(f"  {s['name']}"), fill=True)
            pdf.cell(30, 7, _safe_str(s["type"].upper()), fill=True, align="C")
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(35, 7, f"${s['revenue']:,.0f}  ", fill=True, align="R")
            pdf.cell(35, 7, f"{s['roas']:.2f}x  ", fill=True, align="R")
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(30, 7, _safe_str(s["date"]), fill=True, align="C", ln=True)

    # --- AI Conversation Summary ---
    chat_msgs = _get_chat_history(user_id)
    if chat_msgs:
        if pdf.get_y() > 200:
            pdf.add_page()
            pdf.set_y(20)
        section_header("AI MARKETING CONSULTANT CONVERSATION LOG")
        for msg in chat_msgs[-8:]:  # Include last 8 messages
            role_label = "USER: " if msg["role"] == "user" else "AI CONSULTANT: "
            # Set bold formatting for the speaker label
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(34, 211, 238) if msg["role"] == "user" else pdf.set_text_color(245, 158, 11)
            pdf.write(5, role_label)
            # Normal formatting for the message body
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(50, 60, 70)
            pdf.write(5, _safe_str(msg["message"] + "\n\n"))
        pdf.ln(2)

    # --- Footer ---
    pdf.set_y(-20)
    pdf.set_fill_color(15, 23, 42)
    pdf.rect(0, pdf.get_y(), 210, 20, style="F")
    pdf.set_text_color(130, 140, 160)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(0, 8, _safe_str(f"  ForecastIQ Enterprise Report | {now} | CONFIDENTIAL"), align="L")

    pdf.output(str(filepath))


# ---------------------------------------------------------------------------
# Excel Generator (openpyxl)
# ---------------------------------------------------------------------------

def _generate_xlsx(filepath: Path, dashboard: dict, recs: dict, model_info: dict) -> None:
    """Generate an Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    summary = dashboard.get("summary", {})
    channels = dashboard.get("channel_spend_distribution", [])
    monthly = dashboard.get("monthly_revenue", [])
    recs_list = recs.get("recommendations", [])
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = openpyxl.Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Executive Summary"
    dark_fill = PatternFill("solid", fgColor="0F172A")
    teal_fill = PatternFill("solid", fgColor="22D3EE")
    light_fill = PatternFill("solid", fgColor="F1F5F9")
    head_font = Font(bold=True, color="FFFFFF", size=12)
    label_font = Font(bold=True, color="0F172A", size=10)
    val_font = Font(color="0F172A", size=10)
    teal_font = Font(bold=True, color="0F172A", size=10)

    ws["A1"] = "ForecastIQ — Enterprise Performance Report"
    ws["A1"].font = Font(bold=True, size=16, color="22D3EE")
    ws["A1"].fill = dark_fill
    ws["B1"].fill = dark_fill
    ws["C1"].fill = dark_fill
    ws["A2"] = f"Generated: {now}"
    ws["A2"].font = Font(italic=True, size=9, color="94A3B8")
    ws["A2"].fill = dark_fill
    ws["B2"].fill = dark_fill

    ws.append([])
    ws.append(["EXECUTIVE SUMMARY"])
    ws[ws.max_row][0].font = Font(bold=True, size=12, color="FFFFFF")
    ws[ws.max_row][0].fill = PatternFill("solid", fgColor="1E293B")

    kv_rows = [
        ("Total Revenue", f"${summary.get('total_revenue', 0):,.2f}"),
        ("Total Spend", f"${summary.get('total_spend', 0):,.2f}"),
        ("Overall Profit", f"${summary.get('overall_profit', 0):,.2f}"),
        ("Profit Margin", f"{summary.get('profit_margin', 0):.2f}%"),
        ("Average ROAS", f"{summary.get('average_roas', 0):.2f}x"),
        ("Average CTR", f"{summary.get('average_ctr', 0):.2f}%"),
        ("Avg Daily Revenue", f"${summary.get('average_daily_revenue', 0):,.2f}"),
        ("Avg Daily Spend", f"${summary.get('average_daily_spend', 0):,.2f}"),
    ]
    for i, (k, v) in enumerate(kv_rows):
        ws.append([k, v])
        row_num = ws.max_row
        ws.cell(row_num, 1).font = label_font
        ws.cell(row_num, 2).font = Font(bold=True, color="22D3EE", size=10)
        if i % 2 == 0:
            ws.cell(row_num, 1).fill = light_fill
            ws.cell(row_num, 2).fill = light_fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    # ---- Channel sheet ----
    if channels:
        ws2 = wb.create_sheet("Channel Performance")
        ws2.append(["Channel", "Spend ($)", "Share (%)", "Est. Revenue"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E293B")
            cell.alignment = Alignment(horizontal="center")
        total_rev = summary.get("total_revenue", 0)
        for i, ch in enumerate(channels):
            pct = ch.get("percentage", 0)
            est_rev = round(total_rev * pct / 100, 2)
            ws2.append([
                ch.get("channel", ""),
                round(ch.get("spend", 0), 2),
                round(pct, 2),
                est_rev
            ])
            if i % 2 == 0:
                for cell in ws2[ws2.max_row]:
                    cell.fill = light_fill
        for col in ["A", "B", "C", "D"]:
            ws2.column_dimensions[col].width = 22

    # ---- Monthly Trend sheet ----
    if monthly:
        ws3 = wb.create_sheet("Monthly Revenue")
        ws3.append(["Month", "Revenue ($)"])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E293B")
        for m in monthly:
            ws3.append([m.get("month", ""), round(m.get("revenue", 0), 2)])
        for col in ["A", "B"]:
            ws3.column_dimensions[col].width = 18

    # ---- Recommendations sheet ----
    if recs_list:
        ws4 = wb.create_sheet("AI Recommendations")
        ws4.append(["#", "Priority", "Title", "Description", "Category"])
        for cell in ws4[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E293B")
        for i, rec in enumerate(recs_list, 1):
            ws4.append([
                i,
                rec.get("priority", ""),
                rec.get("title", ""),
                (rec.get("description") or "")[:300],
                rec.get("category", ""),
            ])
            if i % 2 == 0:
                for cell in ws4[ws4.max_row]:
                    cell.fill = light_fill
        ws4.column_dimensions["A"].width = 5
        ws4.column_dimensions["B"].width = 12
        ws4.column_dimensions["C"].width = 35
        ws4.column_dimensions["D"].width = 60
        ws4.column_dimensions["E"].width = 18

    wb.save(str(filepath))


# ---------------------------------------------------------------------------
# CSV Generator
# ---------------------------------------------------------------------------

def _generate_csv(filepath: Path, dashboard: dict, recs: dict, model_info: dict) -> None:
    """Generate a CSV file from key metrics."""
    summary = dashboard.get("summary", {})
    channels = dashboard.get("channel_spend_distribution", [])
    monthly = dashboard.get("monthly_revenue", [])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        writer.writerow(["ForecastIQ Performance Report"])
        writer.writerow(["Generated", now])
        writer.writerow([])

        writer.writerow(["EXECUTIVE SUMMARY"])
        writer.writerow(["Metric", "Value"])
        for k, v in [
            ("Total Revenue", summary.get("total_revenue", 0)),
            ("Total Spend", summary.get("total_spend", 0)),
            ("Overall Profit", summary.get("overall_profit", 0)),
            ("Profit Margin (%)", summary.get("profit_margin", 0)),
            ("Average ROAS", summary.get("average_roas", 0)),
            ("Average CTR (%)", summary.get("average_ctr", 0)),
        ]:
            writer.writerow([k, v])

        writer.writerow([])
        writer.writerow(["CHANNEL PERFORMANCE"])
        writer.writerow(["Channel", "Spend ($)", "Share (%)"])
        for ch in channels:
            writer.writerow([ch.get("channel", ""), ch.get("spend", 0), ch.get("percentage", 0)])

        writer.writerow([])
        writer.writerow(["MONTHLY REVENUE"])
        writer.writerow(["Month", "Revenue ($)"])
        for m in monthly:
            writer.writerow([m.get("month", ""), m.get("revenue", 0)])

        if model_info:
            writer.writerow([])
            writer.writerow(["MODEL PERFORMANCE"])
            writer.writerow(["Model", model_info.get("model_name", "LightGBM")])
            r2 = model_info.get("r2_score", model_info.get("test_r2", 0))
            if r2:
                writer.writerow(["R² Score", r2])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_report(user_id: int, report_type: str, scenario_name: str, paths) -> dict[str, Any]:
    """
    Generate a report file for the authenticated user.
    Returns metadata dict for the reports index.
    """
    reports_dir = _get_user_reports_dir(user_id)
    dashboard = _load_dashboard_data(paths)
    recs = _load_recommendations(paths)
    model_info = _load_model_info(paths)

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    uid = uuid.uuid4().hex[:6].upper()

    # Determine extension
    ext_map = {"pdf": "pdf", "xlsx": "xlsx", "csv": "csv"}
    ext = ext_map.get(report_type, "pdf")
    filename = f"ForecastIQ_Report_{now_str}_{uid}.{ext}"
    filepath = reports_dir / filename

    if report_type == "xlsx":
        _generate_xlsx(filepath, dashboard, recs, model_info)
    elif report_type == "csv":
        _generate_csv(filepath, dashboard, recs, model_info)
    else:
        # Default: PDF
        _generate_pdf(filepath, dashboard, recs, model_info, user_id)

    size_bytes = filepath.stat().st_size if filepath.exists() else 0
    size_str = f"{size_bytes / 1024:.1f} KB" if size_bytes < 1024 * 1024 else f"{size_bytes / 1024 / 1024:.1f} MB"

    meta = {
        "id": uid,
        "name": scenario_name or "Baseline Forecast Scenario",
        "filename": filename,
        "filepath": str(filepath),
        "report_type": report_type,
        "file_size": size_str,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "generated_at": datetime.now().isoformat()
    }

    index_file = reports_dir / "reports_index.json"
    existing = []
    if index_file.exists():
        try:
            with open(index_file, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            existing = []
    existing.insert(0, meta)
    with open(index_file, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2)

    return meta


def get_reports_list(user_id: int, search: Optional[str] = None, report_type: Optional[str] = None, page: int = 1, limit: int = 10) -> dict:
    """Retrieve paginated, filtered list of reports for the authenticated user."""
    reports_dir = _get_user_reports_dir(user_id)
    index_file = reports_dir / "reports_index.json"

    items = []
    if index_file.exists():
        try:
            with open(index_file, "r", encoding="utf-8") as f:
                items = json.load(f)
        except Exception:
            items = []

    # Filter
    if search:
        s = search.lower()
        items = [r for r in items if s in r.get("filename", "").lower()]
    if report_type and report_type != "all":
        items = [r for r in items if r.get("report_type") == report_type]

    total = len(items)
    start = (page - 1) * limit
    paginated = items[start:start + limit]
    pages = max(1, (total + limit - 1) // limit)

    return {
        "items": paginated,
        "total": total,
        "page": page,
        "pages": pages
    }


def delete_report(user_id: int, report_id: str) -> bool:
    """Delete a report and remove it from the index."""
    reports_dir = _get_user_reports_dir(user_id)
    index_file = reports_dir / "reports_index.json"

    if not index_file.exists():
        return False

    try:
        with open(index_file, "r", encoding="utf-8") as f:
            items = json.load(f)

        to_delete = next((r for r in items if r["id"] == report_id), None)
        if not to_delete:
            return False

        fp = Path(to_delete.get("filepath", ""))
        if fp.exists():
            fp.unlink()

        items = [r for r in items if r["id"] != report_id]
        with open(index_file, "w", encoding="utf-8") as f:
            json.dump(items, f, indent=2)
        return True
    except Exception as exc:
        logger.error("Delete report error: %s", exc)
        return False


def get_report_filepath(user_id: int, report_id: str) -> Optional[Path]:
    """Return the file path of a specific report for the authenticated user."""
    reports_dir = _get_user_reports_dir(user_id)
    index_file = reports_dir / "reports_index.json"
    if not index_file.exists():
        return None
    try:
        with open(index_file, "r", encoding="utf-8") as f:
            items = json.load(f)
        report = next((r for r in items if r["id"] == report_id), None)
        if report:
            fp = Path(report.get("filepath", ""))
            if fp.exists():
                return fp
    except Exception as exc:
        logger.error("Get filepath error: %s", exc)
    return None
